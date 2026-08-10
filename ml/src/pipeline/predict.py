"""
predict.py -- score an input file and produce the three target columns.
======================================================================

    Vessel_Training_Input_Sample.xlsx          trained_models/uc1_m3_tat_*.pkl
                    |                                     |
                    +------------------+------------------+
                                       v
                                 [ predict.py ]
                                       |
                                       v
             out/predictions.xlsx  +  out/predictions.json
                                    +  out/predictions_dashboard.json
                       ETB | TAT | ETD  (+ P10/P90 bands and drivers)

``predictions.json`` is the audit trail -- every formula and every substituted
value behind each number. ``predictions_dashboard.json`` is the same
predictions with only the fields a UI draws, and a glossary of its own keys.

This is the inference half of the pair. It loads a model that was trained
*once* by ``train_tat_model.py`` and frozen to disk; it never fits anything.
Two runs against the same input and the same artefact give byte-identical
numbers, which is what makes a prediction quotable.

HOW THE THREE TARGETS ARE PRODUCED
----------------------------------
The README sheet in the sample workbook asks for ETB, TAT and ETD. Only one of
them is a learned quantity; the other two are derived from it by arithmetic that
is spelled out per row in the JSON output.

    TAT   <- UC1-M3, the trained model.   TAT = ATA -> ATD, in hours.
             Returned as p10 / p50 / p90; p50 is the headline.

    ETB   <- ATA + pre-berth wait.
             The wait comes from whichever ``--wait-model`` you choose:

               optimiser  (default)  UC1-M5 assigns berths under LOA, draft,
                                     non-overlap and tidal constraints; the
                                     wait is the gap between the requested
                                     start and the assigned start.
               queue                 UC1-M4's waiting-time distribution (p50 by
                                     default, or p90 with --wait-percentile 90),
                                     measured from berthing logs.
               none                  wait = 0; ETB = ATA.

    ETD   <- ATA + TAT.

    berth stay = TAT - wait, i.e. ETD - ETB.

WHERE THIS CAN MISLEAD YOU, STATED UP FRONT
-------------------------------------------
1. ``--wait-model optimiser`` computes contention **within your input file
   only.** If the file is a three-row sample rather than the port's full
   arrival list, almost nothing competes for a berth and the wait it reports is
   a *lower bound*, not a forecast. Every affected row is flagged
   ``WAIT_IS_LOWER_BOUND`` and the summary says so. For a realistic wait on a
   partial file, use ``--wait-model queue``.

2. The residual berth stay can come out implausibly short when the wait
   estimate and the TAT estimate disagree -- e.g. a queue-model wait of 20 h
   against a predicted TAT of 24 h leaves only 4 h alongside for a 4,800 TEU
   parcel. Those rows are reconciled: berth stay is floored at
   ``MIN_BERTH_STAY_H``, ETD is pushed out to keep ETB + stay = ETD, and the row
   is flagged ``ETD_RECONCILED``. The unreconciled TAT is kept alongside so you
   can see exactly what moved.

3. When an ML engine supplies p50, the per-row driver list explains the
   *additive surrogate*, not the gradient-boosted model's internal splits. The
   output column ``Attribution_Source`` says which, on every row. Do not present
   a LightGBM number with an additive explanation and call it the model's
   reasoning.

CLI
---
    python run.py predict --input data/input/Vessel_Training_Input_Sample.xlsx
    python run.py predict --input data.xlsx --artifact trained_models/uc1_m3_tat_lightgbm_v1.2.0.pkl
    python run.py predict --input data.xlsx --wait-model queue --wait-percentile 90
    python run.py predict --input data.xlsx --out out/ --format xlsx
    python run.py predict --selftest

With no ``--artifact`` the newest ``trained_models/uc1_m3_tat*.pkl`` is used. If there
is none, the run falls back to the zero-dependency additive model and says so
loudly -- a fallback that is invisible is a fallback that gets quoted.
"""

from __future__ import annotations

import argparse
import csv
import json
import math
import os
import sys
from dataclasses import dataclass, field, asdict
from datetime import datetime, timedelta, timezone
from typing import Any, Dict, List, Optional, Sequence, Tuple

import jnpa_paths

jnpa_paths.ensure_on_syspath()

import jnpa_input as jio               # noqa: E402
import train_tat_model as trainer      # noqa: E402
import uc1_m3_tat_predict as m3        # noqa: E402

MODULE_ID: str = "UC1-PREDICT"
MODULE_VERSION: str = "predict-v1.0.0"

MIN_BERTH_STAY_H: float = 6.0
"""
Floor on the residual berth stay. A container call cannot discharge and load in
under a few hours; when the arithmetic implies less, the wait and TAT estimates
have disagreed and the row is reconciled and flagged rather than published as-is.
"""

PLAN_CLUSTER_GAP_H: float = 72.0
"""
Requests separated by more than this are put in different berth-planning
clusters. Vessels arriving three weeks apart do not compete for a berth, so
optimising them in one 72 h horizon would invent contention that does not exist.
"""

JNPA_ANCHOR_TAT_H: float = 43.92     # 1.83 days, JNPA published reference
JNPA_ANCHOR_BERTH_STAY_H: float = 23.28   # 0.97 days
JNPA_ANCHOR_WAIT_H: float = JNPA_ANCHOR_TAT_H - JNPA_ANCHOR_BERTH_STAY_H


def _ist(dt: Optional[datetime]) -> str:
    if dt is None:
        return ""
    return dt.astimezone(jio.IST).strftime("%Y-%m-%d %H:%M")


def _iso(dt: Optional[datetime]) -> Optional[str]:
    return dt.isoformat() if dt else None


# ---------------------------------------------------------------------------
# result record
# ---------------------------------------------------------------------------


@dataclass
class Prediction:
    """One row of output: the three targets plus everything needed to defend them."""

    row: int
    call_id: str
    vessel_name: str
    imo: str
    voyage: str
    terminal: str
    requested_berth: str
    assigned_berth: str

    ata_utc: Optional[datetime]
    etb_utc: Optional[datetime]
    etd_utc: Optional[datetime]

    tat_hours: float
    tat_p10_hours: float
    tat_p90_hours: float
    tat_days: float
    tat_hours_unreconciled: float

    wait_hours: float
    berth_stay_hours: float
    wait_source: str

    engine: str
    attribution_source: str
    confidence: str
    band_width_hours: float
    sigma_hours: float
    stressors: Tuple[str, ...]
    drivers: Tuple[Dict[str, Any], ...]
    flags: Tuple[str, ...]
    notes: Tuple[str, ...]
    breakdown: Dict[str, Any] = field(default_factory=dict)

    def as_row(self) -> Dict[str, Any]:
        """The flat, spreadsheet-shaped view."""
        return {
            "Row": self.row,
            "Vessel": self.vessel_name,
            "IMO": self.imo,
            "Voyage": self.voyage,
            "Terminal": self.terminal,
            "Requested_Berth": self.requested_berth,
            "Assigned_Berth": self.assigned_berth,
            "ATA_IST": _ist(self.ata_utc),
            "ETB_IST": _ist(self.etb_utc),
            "ETD_IST": _ist(self.etd_utc),
            "TAT_Hours": round(self.tat_hours, 2),
            "TAT_Days": round(self.tat_days, 3),
            "TAT_P10_Hours": round(self.tat_p10_hours, 2),
            "TAT_P90_Hours": round(self.tat_p90_hours, 2),
            "Wait_Hours": round(self.wait_hours, 2),
            "Berth_Stay_Hours": round(self.berth_stay_hours, 2),
            "Confidence": self.confidence,
            "Band_Width_Hours": round(self.band_width_hours, 2),
            "Engine": self.engine,
            "Attribution_Source": self.attribution_source,
            "Wait_Source": self.wait_source,
            "Top_Driver_1": self._driver(0),
            "Top_Driver_2": self._driver(1),
            "Top_Driver_3": self._driver(2),
            "Stressors": ", ".join(self.stressors),
            "Flags": ", ".join(self.flags),
        }

    def _driver(self, i: int) -> str:
        if i >= len(self.drivers):
            return ""
        d = self.drivers[i]
        return f"{d.get('factor')} {d.get('contribution_h'):+.2f} h ({d.get('share_pct')}%)"

    def as_dict(self) -> Dict[str, Any]:
        out = self.as_row()
        out.update({
            "call_id": self.call_id,
            "ata_utc": _iso(self.ata_utc),
            "etb_utc": _iso(self.etb_utc),
            "etd_utc": _iso(self.etd_utc),
            "tat_hours_unreconciled": round(self.tat_hours_unreconciled, 3),
            "sigma_hours": round(self.sigma_hours, 3),
            "drivers": list(self.drivers),
            "notes": list(self.notes),
            "breakdown": self.breakdown,
        })
        return out


# ---------------------------------------------------------------------------
# model loading
# ---------------------------------------------------------------------------


def _candidate_artifacts(artifact: Optional[str], model_dir: str) -> List[str]:
    """Artefacts to try, newest first. An explicit ``--artifact`` is tried alone."""
    if artifact:
        return [artifact]
    if not os.path.isdir(model_dir):
        return []
    found = [
        os.path.join(model_dir, f) for f in os.listdir(model_dir)
        if f.endswith(".pkl") and f.startswith("uc1_m3_tat")
    ]
    return sorted(found, key=os.path.getmtime, reverse=True)


def _load_one(path: str) -> Tuple[Any, Dict[str, Any]]:
    payload = trainer.load_artifact(path)
    card = payload.get("model_card", {})
    return payload["predictor"], {
        "mode": "TRAINED_ARTIFACT",
        "artifact_path": payload["artifact_path"],
        "artifact_sha256": card.get("artifact_sha256", ""),
        "engine": payload.get("engine", "unknown"),
        "model_version": payload.get("model_version", ""),
        "trained_at_utc": payload.get("created_at_utc", ""),
        "trained_on_days": payload.get("config", {}).get("days"),
        "holdout_mae_hours": card.get("metrics_holdout", {}).get("mae_hours"),
        "holdout_coverage_pct": card.get("metrics_holdout", {}).get("coverage_80_pct"),
        "warning": "",
    }


def resolve_predictor(artifact: Optional[str], model_dir: str) -> Tuple[Any, Dict[str, Any]]:
    """
    Return ``(predictor, provenance)``.

    Order: the explicit ``--artifact``, else every artefact in ``model_dir``
    newest first, else an in-process additive fallback. The provenance dict
    always records which happened and why.

    A LightGBM artefact cannot be unpickled on a machine without LightGBM, so a
    load failure of that kind is survivable: we move to the next artefact -- the
    ``_portable`` additive one exists precisely for this -- and only then fit
    in-process. Two failures are NOT survivable and propagate:

      * ``FileNotFoundError`` when you named an artefact explicitly -- silently
        scoring with a different model than the one you asked for would be worse
        than stopping.
      * a SHA-256 digest mismatch -- the file has changed since it was trained,
        which is an integrity failure, not a compatibility one.
    """
    attempts: List[str] = []
    for path in _candidate_artifacts(artifact, model_dir):
        try:
            predictor, provenance = _load_one(path)
            if attempts:
                provenance["warning"] = (
                    "Fell back to " + os.path.basename(path) + " after: "
                    + "; ".join(attempts)
                )
            return predictor, provenance
        except (FileNotFoundError, ValueError):
            # Missing file you named, wrong schema, or a digest mismatch. These
            # are user-actionable and must not be papered over.
            raise
        except Exception as exc:
            # Typically a missing engine library (LightGBM/sklearn) on this
            # machine. Try the next artefact.
            attempts.append(f"{os.path.basename(path)} -> {type(exc).__name__}: {exc}"[:160])
            continue

    # Nothing loadable. Fit the transparent additive model in-process so the run
    # still produces numbers, and make the downgrade impossible to miss.
    predictor = m3.TATPredictor(engine="additive")
    predictor.fit(m3.generate_synthetic_calls(n_days=60))
    if attempts:
        warning = (
            "No trained artefact could be loaded on this machine ("
            + "; ".join(attempts) + "). Used the transparent additive model fitted "
            "in-process. Retrain here with "
            f"'python run.py train --engine additive --out {model_dir}/'."
        )
    else:
        warning = (
            f"No trained artefact found in {model_dir!r}. Used the transparent "
            f"additive model fitted in-process. Run "
            f"'python run.py train --out {model_dir}/' for the trained model."
        )
    return predictor, {
        "mode": "UNTRAINED_FALLBACK",
        "artifact_path": "",
        "artifact_sha256": "",
        "engine": "additive",
        "model_version": m3.TAT_MODEL_VERSION,
        "trained_at_utc": datetime.now(timezone.utc).isoformat(),
        "trained_on_days": 60,
        "holdout_mae_hours": None,
        "holdout_coverage_pct": None,
        "warning": warning,
    }


# ---------------------------------------------------------------------------
# wait models
# ---------------------------------------------------------------------------


def _cluster_by_time(calls: Sequence[jio.VesselCallInput],
                     gap_hours: float = PLAN_CLUSTER_GAP_H) -> List[List[jio.VesselCallInput]]:
    """Split calls into berth-planning clusters separated by more than ``gap_hours``."""
    dated = sorted((c for c in calls if c.ata_utc), key=lambda c: (c.ata_utc, c.call_id))
    clusters: List[List[jio.VesselCallInput]] = []
    for call in dated:
        if clusters and (call.ata_utc - clusters[-1][-1].ata_utc).total_seconds() / 3600.0 <= gap_hours:
            clusters[-1].append(call)
        else:
            clusters.append([call])
    return clusters


def wait_from_optimiser(calls: Sequence[jio.VesselCallInput],
                        gap_hours: float = PLAN_CLUSTER_GAP_H) -> Dict[str, Dict[str, Any]]:
    """
    Run M5 per planning cluster and return ``{call_id: {...}}``.

    Tidal windows come from M2's scanner over each cluster's span, so the tide
    used for berth planning is the same tide used everywhere else in the run.
    """
    import uc1_m2_tidal_window as m2
    import uc1_m5_berth_optimiser as m5

    out: Dict[str, Dict[str, Any]] = {}
    berths = m5.default_berths()

    for cluster in _cluster_by_time(calls, gap_hours):
        requests = [jio.to_m5_request(c) for c in cluster]
        start = min(r.requested_start_utc for r in requests) - timedelta(hours=6)
        span_h = max(
            72.0,
            (max(r.requested_start_utc for r in requests) - start).total_seconds() / 3600.0
            + max(r.service_hours for r in requests) + 24.0,
        )
        deepest = max(c.draft_m for c in cluster)
        probe = jio.to_m2_vessel(max(cluster, key=lambda c: c.draft_m))
        scan, _ = m2.scan_windows(
            probe, m2.SyntheticTideProvider(), start, m2.DEFAULT_REACHES,
            siltation_m=max(c.siltation_m for c in cluster),
            dredging_m=min(c.dredging_delta_m for c in cluster),
            hours=min(span_h, 240.0),
        )
        windows = [
            m5.TidalWindow(
                window_id=f"TW-{w.window_id:02d}",
                start_utc=w.start, end_utc=w.end,
                max_draft_m=deepest, reach_id=w.binding_reach_id, direction="INBOUND",
            )
            for w in scan.windows
        ]
        plan = m5.optimise(requests, berths, windows, algorithm="auto", tide_policy="soft")
        for a in plan.assignments:
            out[a.request_id] = {
                "wait_hours": a.wait_hours,
                "assigned_berth": a.berth_id,
                "is_berth_shift": a.is_berth_shift,
                "tide_miss": a.tide_miss,
                "start_utc": a.start_utc,
                "rationale": a.rationale,
                "algorithm": plan.algorithm,
                "cluster_size": len(cluster),
                "tidal_windows_available": len(windows),
            }
        for rid in plan.unassigned_request_ids:
            out[rid] = {
                "wait_hours": 0.0, "assigned_berth": "", "is_berth_shift": False,
                "tide_miss": True, "start_utc": None,
                "rationale": "no feasible berth found in this planning cluster",
                "algorithm": plan.algorithm, "cluster_size": len(cluster),
                "tidal_windows_available": len(windows),
            }
    return out


def wait_from_queue(percentile: int = 50) -> Dict[str, Any]:
    """
    Waiting time from M4's berthing-log distribution -- a port-level statistic,
    not a per-vessel one.

    The real DSR-derived log is tried first, but it cannot answer this question:
    Daily Status Report section (H) records *berthed on* and *expected
    completion*, never an arrival at anchorage. Waiting time is defined as
    ``actual_atb - actual_ata``, so with no ATA every record is dropped and the
    distribution is empty. When that happens we fall back to the calibrated
    synthetic berthing log and say exactly why in ``fallback_reason`` -- an
    empty real distribution must not silently become a NaN in a berth plan.
    """
    import uc1_m4_berth_utilisation as m4

    end = datetime.now(timezone.utc)
    start = end - timedelta(days=120)
    records, _berths, source = m4.load_records_with_fallback(start, end)
    stats = m4.waiting_time_distribution(records)

    fallback_reason = ""
    if stats.n == 0 or not math.isfinite(stats.p50_hours):
        reasons = {d.get("reason", "unknown") for d in stats.drop_reasons}
        fallback_reason = (
            f"{source} yielded no usable waiting times "
            f"({stats.n_dropped} records dropped; reasons: {', '.join(sorted(reasons)) or 'none'}). "
            f"The DSR corpus records berthing and completion but never an arrival at "
            f"anchorage, so '{stats.definition}' is not computable from it. "
            f"Fell back to the calibrated synthetic berthing log."
        )
        loader = m4.SyntheticBerthingLogLoader()
        records = loader.load_records(start, end)
        source = f"{loader.source_id} (fallback)"
        stats = m4.waiting_time_distribution(records)

    value = {10: stats.p10_hours, 50: stats.p50_hours, 90: stats.p90_hours}.get(
        percentile, stats.p50_hours
    )
    if not math.isfinite(value):
        raise ValueError(
            "no waiting-time distribution could be built from any source; "
            "use --wait-model optimiser or --wait-model none"
        )
    return {
        "wait_hours": float(value),
        "percentile": percentile,
        "n": stats.n,
        "n_dropped": stats.n_dropped,
        "data_source": source,
        "definition": stats.definition,
        "p50": round(stats.p50_hours, 3),
        "p90": round(stats.p90_hours, 3),
        "fallback_reason": fallback_reason,
    }


# ---------------------------------------------------------------------------
# the prediction pass
# ---------------------------------------------------------------------------


def predict_batch(
    batch: jio.InputBatch,
    predictor: Any,
    *,
    wait_model: str = "optimiser",
    wait_percentile: int = 50,
    cluster_gap_h: float = PLAN_CLUSTER_GAP_H,
) -> Tuple[List[Prediction], Dict[str, Any]]:
    """Score every valid row and return ``(predictions, run_context)``."""
    calls = batch.valid_rows
    context: Dict[str, Any] = {"wait_model": wait_model}

    optimiser_waits: Dict[str, Dict[str, Any]] = {}
    queue_stats: Dict[str, Any] = {}
    if wait_model == "optimiser":
        optimiser_waits = wait_from_optimiser(calls, cluster_gap_h)
        context["optimiser"] = {
            "clusters": len(_cluster_by_time(calls, cluster_gap_h)),
            "assigned": sum(1 for v in optimiser_waits.values() if v["assigned_berth"]),
            "algorithms": sorted({v["algorithm"] for v in optimiser_waits.values()}),
        }
    elif wait_model == "queue":
        queue_stats = wait_from_queue(wait_percentile)
        context["queue"] = queue_stats

    predictions: List[Prediction] = []
    for call in calls:
        flags: List[str] = []
        notes: List[str] = []

        features = jio.to_m3_features(call)
        pred = predictor.predict(features)

        tat_p50 = float(pred.p50_hours)
        tat_p10 = float(pred.p10_hours)
        tat_p90 = float(pred.p90_hours)

        # --- pre-berth wait -------------------------------------------------
        assigned_berth = call.requested_berth
        if wait_model == "optimiser":
            info = optimiser_waits.get(call.call_id, {})
            wait_h = float(info.get("wait_hours", 0.0))
            assigned_berth = info.get("assigned_berth") or call.requested_berth
            wait_source = f"M5_OPTIMISER({info.get('algorithm', 'greedy')})"
            if info.get("rationale"):
                notes.append(str(info["rationale"]))
            if info.get("is_berth_shift"):
                flags.append("BERTH_SHIFT")
            if info.get("tide_miss"):
                flags.append("TIDE_MISS")
            if int(info.get("cluster_size", 1)) <= 2:
                flags.append("WAIT_IS_LOWER_BOUND")
                notes.append(
                    f"only {info.get('cluster_size', 1)} vessel(s) in this planning "
                    f"cluster, so berth contention is near zero; the wait is a lower "
                    f"bound. Use --wait-model queue for a distribution-based estimate."
                )
        elif wait_model == "queue":
            wait_h = float(queue_stats["wait_hours"])
            wait_source = (f"M4_QUEUE_P{wait_percentile}"
                           f"({queue_stats['data_source']}, n={queue_stats['n']})")
            notes.append(
                f"port-level waiting-time p{wait_percentile} = {wait_h:.2f} h applied to "
                f"every vessel; this is a distribution statistic, not a per-vessel forecast"
            )
        else:
            wait_h, wait_source = 0.0, "NONE(ETB=ATA)"

        # --- assemble the three targets -------------------------------------
        ata = call.ata_utc
        tat_unreconciled = tat_p50
        berth_stay = tat_p50 - wait_h
        if berth_stay < MIN_BERTH_STAY_H:
            berth_stay = MIN_BERTH_STAY_H
            tat_p50 = wait_h + berth_stay
            shift = tat_p50 - tat_unreconciled
            tat_p90 = max(tat_p90, tat_p50)
            tat_p10 = min(tat_p10, tat_p50)
            flags.append("ETD_RECONCILED")
            notes.append(
                f"predicted TAT {tat_unreconciled:.2f} h minus wait {wait_h:.2f} h left only "
                f"{tat_unreconciled - wait_h:.2f} h alongside; berth stay floored at "
                f"{MIN_BERTH_STAY_H:.0f} h and ETD pushed out {shift:.2f} h"
            )

        etb = ata + timedelta(hours=wait_h) if ata else None
        etd = ata + timedelta(hours=tat_p50) if ata else None

        # --- explainability -------------------------------------------------
        contributions = pred.breakdown.get("contributions", [])
        drivers = tuple(
            sorted(contributions, key=lambda d: abs(float(d.get("contribution_h", 0.0))),
                   reverse=True)[:5]
        )
        attribution = pred.breakdown.get("attribution_source", "additive")
        if pred.engine != "additive":
            notes.append(
                f"p50 came from '{pred.engine}'; the driver list explains the additive "
                f"surrogate, not that model's internal splits"
            )
        if pred.quantile_crossing_corrected:
            flags.append("QUANTILE_CROSSING_CORRECTED")
        if pred.clamped_at_min:
            flags.append("CLAMPED_AT_MIN_TAT")
        if call.tide_source == "SYNTHETIC_HARMONIC_v1":
            flags.append("TIDE_SYNTHETIC")
        if call.queue_source == "DERIVED_FROM_OCCUPANCY":
            flags.append("QUEUE_DERIVED")
        if call.cranes_available:
            notes.append(
                f"Cranes_Available={call.cranes_available} was read but is NOT an M3 "
                f"feature; adding it requires editing FEATURE_COLUMNS and retraining"
            )

        band = tat_p90 - tat_p10
        confidence = "HIGH" if band <= 8.0 else ("MEDIUM" if band <= 16.0 else "LOW")

        predictions.append(Prediction(
            row=call.row,
            call_id=call.call_id,
            vessel_name=call.vessel_name,
            imo=call.imo,
            voyage=call.voyage,
            terminal=call.terminal,
            requested_berth=call.requested_berth,
            assigned_berth=assigned_berth,
            ata_utc=ata,
            etb_utc=etb,
            etd_utc=etd,
            tat_hours=tat_p50,
            tat_p10_hours=tat_p10,
            tat_p90_hours=tat_p90,
            tat_days=tat_p50 / 24.0,
            tat_hours_unreconciled=tat_unreconciled,
            wait_hours=wait_h,
            berth_stay_hours=berth_stay,
            wait_source=wait_source,
            engine=pred.engine,
            attribution_source=attribution,
            confidence=confidence,
            band_width_hours=band,
            sigma_hours=float(pred.sigma_hours),
            stressors=tuple(pred.stressors_active),
            drivers=drivers,
            flags=tuple(dict.fromkeys(flags)),
            notes=tuple(notes),
            breakdown={
                "targets": {
                    "TAT": {
                        "formula": "TAT_hours = M3.predict(features).p50",
                        "value_h": round(tat_p50, 3),
                        "p10_h": round(tat_p10, 3),
                        "p90_h": round(tat_p90, 3),
                        "engine": pred.engine,
                    },
                    "ETB": {
                        "formula": "ETB = ATA + wait_hours",
                        "substitution": f"{_ist(ata)} IST + {wait_h:.2f} h = {_ist(etb)} IST",
                        "wait_source": wait_source,
                    },
                    "ETD": {
                        "formula": "ETD = ATA + TAT_hours",
                        "substitution": f"{_ist(ata)} IST + {tat_p50:.2f} h = {_ist(etd)} IST",
                    },
                    "berth_stay": {
                        "formula": "berth_stay = TAT - wait = ETD - ETB",
                        "substitution": f"{tat_p50:.2f} - {wait_h:.2f} = {berth_stay:.2f} h",
                        "floor_h": MIN_BERTH_STAY_H,
                        "reconciled": "ETD_RECONCILED" in flags,
                    },
                },
                "anchors": {
                    "jnpa_mean_tat_h": JNPA_ANCHOR_TAT_H,
                    "jnpa_mean_berth_stay_h": JNPA_ANCHOR_BERTH_STAY_H,
                    "jnpa_implied_mean_wait_h": round(JNPA_ANCHOR_WAIT_H, 2),
                    "note": "public JNPA performance reference, for orientation only",
                },
                "input_provenance": {
                    "tide_source": call.tide_source,
                    "depth_source": call.depth_source,
                    "queue_source": call.queue_source,
                    "speed_source": call.speed_source,
                },
                "m3_breakdown": pred.breakdown,
            },
        ))

    return predictions, context


# ---------------------------------------------------------------------------
# output writers
# ---------------------------------------------------------------------------

OUTPUT_COLUMNS: Tuple[str, ...] = (
    "Row", "Vessel", "IMO", "Voyage", "Terminal", "Requested_Berth", "Assigned_Berth",
    "ATA_IST", "ETB_IST", "ETD_IST",
    "TAT_Hours", "TAT_Days", "TAT_P10_Hours", "TAT_P90_Hours",
    "Wait_Hours", "Berth_Stay_Hours",
    "Confidence", "Band_Width_Hours", "Engine", "Attribution_Source", "Wait_Source",
    "Top_Driver_1", "Top_Driver_2", "Top_Driver_3", "Stressors", "Flags",
)


def write_csv(predictions: Sequence[Prediction], path: str) -> str:
    os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
    with open(path, "w", encoding="utf-8", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=list(OUTPUT_COLUMNS), extrasaction="ignore")
        w.writeheader()
        for p in predictions:
            w.writerow(p.as_row())
    return path


def write_xlsx(predictions: Sequence[Prediction], path: str,
               context: Dict[str, Any], provenance: Dict[str, Any]) -> str:
    """Three sheets: Predictions, Drivers, Run_Info. Falls back to CSV without openpyxl."""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except Exception:
        return write_csv(predictions, os.path.splitext(path)[0] + ".csv")

    os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Predictions"
    ws.append(list(OUTPUT_COLUMNS))
    header_fill = PatternFill("solid", fgColor="1F4E78")
    target_fill = PatternFill("solid", fgColor="C6E0B4")
    for i, name in enumerate(OUTPUT_COLUMNS, start=1):
        c = ws.cell(row=1, column=i)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = max(
            11, min(30, len(name) + 3)
        )
    target_cols = {OUTPUT_COLUMNS.index(n) + 1 for n in ("ETB_IST", "ETD_IST", "TAT_Hours")}
    for p in predictions:
        row = p.as_row()
        ws.append([row.get(k, "") for k in OUTPUT_COLUMNS])
        for col in target_cols:
            ws.cell(row=ws.max_row, column=col).fill = target_fill
    ws.freeze_panes = "A2"

    dv = wb.create_sheet("Drivers")
    dv.append(["Row", "Vessel", "Rank", "Factor", "Input", "Coefficient",
               "Contribution_h", "Share_pct", "Direction"])
    for c in dv[1]:
        c.font = Font(bold=True)
    for p in predictions:
        for rank, d in enumerate(p.drivers, start=1):
            dv.append([p.row, p.vessel_name, rank, d.get("factor"), d.get("input"),
                       d.get("coefficient"), d.get("contribution_h"),
                       d.get("share_pct"), d.get("direction")])
    for col, width in (("A", 6), ("B", 26), ("C", 6), ("D", 24), ("E", 12),
                       ("F", 26), ("G", 15), ("H", 11), ("I", 12)):
        dv.column_dimensions[col].width = width

    info = wb.create_sheet("Run_Info")
    info.append(["Key", "Value"])
    info["A1"].font = Font(bold=True)
    info["B1"].font = Font(bold=True)
    rows: List[Tuple[str, Any]] = [
        ("generated_by", MODULE_VERSION),
        ("generated_at_utc", datetime.now(timezone.utc).isoformat()),
        ("", ""),
        ("model_mode", provenance.get("mode")),
        ("model_engine", provenance.get("engine")),
        ("model_version", provenance.get("model_version")),
        ("artifact_path", provenance.get("artifact_path")),
        ("artifact_sha256", provenance.get("artifact_sha256")),
        ("trained_at_utc", provenance.get("trained_at_utc")),
        ("trained_on_days", provenance.get("trained_on_days")),
        ("holdout_mae_hours", provenance.get("holdout_mae_hours")),
        ("holdout_coverage_pct", provenance.get("holdout_coverage_pct")),
        ("model_warning", provenance.get("warning")),
        ("", ""),
        ("wait_model", context.get("wait_model")),
    ]
    for key in ("optimiser", "queue"):
        if key in context:
            for k, v in context[key].items():
                rows.append((f"{key}.{k}", v))
    rows += [
        ("", ""),
        ("target_TAT", "M3 trained model, p50 of the predicted ATA->ATD hours"),
        ("target_ETB", "ATA + pre-berth wait (see wait_model)"),
        ("target_ETD", "ATA + TAT"),
        ("berth_stay", f"TAT - wait, floored at {MIN_BERTH_STAY_H:g} h (flag ETD_RECONCILED)"),
        ("", ""),
        ("caveat_1", "Cranes_Available is read but is not an M3 feature; it needs a retrain"),
        ("caveat_2", "DUKC_Status in the input is M1's OUTPUT; used only to score M1"),
        ("caveat_3", "Tide height is SYNTHETIC unless a Tide_Height_m column was supplied"),
    ]
    for k, v in rows:
        # openpyxl accepts only scalars; lists/dicts are rendered for the reader.
        if isinstance(v, (list, tuple, set)):
            v = ", ".join(str(x) for x in v)
        elif isinstance(v, dict):
            v = json.dumps(v, default=str)
        info.append([k, v])
    info.column_dimensions["A"].width = 26
    info.column_dimensions["B"].width = 96

    wb.save(path)
    return path


def write_json(predictions: Sequence[Prediction], path: str,
               batch: jio.InputBatch, context: Dict[str, Any],
               provenance: Dict[str, Any]) -> str:
    os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
    payload = {
        "generated_by": MODULE_VERSION,
        "generated_at_utc": datetime.now(timezone.utc).isoformat(),
        "input": batch.summary(),
        "model": provenance,
        "run_context": context,
        "target_definitions": {
            "TAT": "M3 trained model; p50 of predicted ATA->ATD hours",
            "ETB": "ATA + pre-berth wait",
            "ETD": "ATA + TAT",
            "berth_stay": f"TAT - wait, floored at {MIN_BERTH_STAY_H} h",
        },
        "predictions": [p.as_dict() for p in predictions],
    }
    with open(path, "w", encoding="utf-8") as fh:
        json.dump(payload, fh, indent=2, default=str)

    # Companion file: the same predictions with only the fields a dashboard
    # draws. The payload above exists to be audited -- it carries the formula and
    # the substituted values behind every number, which is what makes the figure
    # defensible and also what makes it unreadable at a glance.
    import dashboard_json

    dashboard_json.write(
        dashboard_json.build(
            [{
                "model_id": "UC1-M3",
                "ok": True,
                "rows": [p.as_row() for p in predictions],
                "summary": {
                    "vessels": len(predictions),
                    "engine": provenance.get("engine"),
                    "model_mode": provenance.get("mode"),
                    "holdout_mae_hours": provenance.get("holdout_mae_hours"),
                    "holdout_coverage_pct": provenance.get("holdout_coverage_pct"),
                    "mean_tat_hours": (
                        round(sum(p.tat_hours for p in predictions) / len(predictions), 2)
                        if predictions else 0
                    ),
                },
            }],
            batch,
            {"wait_model": context.get("wait_model")},
            generated_at_utc=payload["generated_at_utc"],
            full_detail_file=jnpa_paths.relative(path),
        ),
        os.path.splitext(path)[0] + "_dashboard.json",
    )
    return path


def dashboard_path_for(json_path: str) -> str:
    """Where write_json puts the dashboard companion of ``json_path``."""
    return os.path.splitext(json_path)[0] + "_dashboard.json"


# ---------------------------------------------------------------------------
# console report
# ---------------------------------------------------------------------------


def _print_report(predictions: Sequence[Prediction], batch: jio.InputBatch,
                  context: Dict[str, Any], provenance: Dict[str, Any]) -> None:
    print("=" * 112)
    print(f"  JNPA UC-1 PREDICTION  |  {MODULE_VERSION}  |  targets: ETB, TAT, ETD")
    print("=" * 112)
    print(f"  input   : {batch.source_file}")
    print(f"            {len(batch.valid_rows)}/{len(batch.rows)} rows valid, "
          f"{batch.error_count} errors, {batch.warn_count} warnings")
    print(f"  model   : {provenance['mode']}  engine={provenance['engine']}  "
          f"version={provenance['model_version']}")
    if provenance.get("artifact_path"):
        print(f"            {provenance['artifact_path']}")
        print(f"            sha256 {provenance['artifact_sha256'][:16]}...  "
              f"trained {str(provenance.get('trained_at_utc'))[:19]}  "
              f"on {provenance.get('trained_on_days')} days")
        mae, cov = provenance.get("holdout_mae_hours"), provenance.get("holdout_coverage_pct")
        if mae is not None:
            print(f"            held-out MAE {mae:.2f} h, 80% band coverage {cov:.1f}%")
    if provenance.get("warning"):
        print(f"\n  !! {provenance['warning']}\n")
    print(f"  wait    : {context.get('wait_model')}")
    print()

    hdr = (f"  {'row':>3}  {'vessel':<23} {'ATA (IST)':<16} {'ETB (IST)':<16} "
           f"{'ETD (IST)':<16} {'TAT h':>7} {'p10':>6} {'p90':>6} "
           f"{'wait':>6} {'stay':>6}  {'conf':<7} berth")
    print("  TARGETS")
    print(hdr)
    print("  " + "-" * (len(hdr) - 2))
    for p in predictions:
        print(f"  {p.row:>3}  {p.vessel_name[:23]:<23} {_ist(p.ata_utc):<16} "
              f"{_ist(p.etb_utc):<16} {_ist(p.etd_utc):<16} "
              f"{p.tat_hours:>7.2f} {p.tat_p10_hours:>6.1f} {p.tat_p90_hours:>6.1f} "
              f"{p.wait_hours:>6.2f} {p.berth_stay_hours:>6.2f}  "
              f"{p.confidence:<7} {p.assigned_berth}")
    print()

    if predictions:
        mean_tat = sum(p.tat_hours for p in predictions) / len(predictions)
        mean_stay = sum(p.berth_stay_hours for p in predictions) / len(predictions)
        mean_wait = sum(p.wait_hours for p in predictions) / len(predictions)
        print("  BATCH MEANS vs JNPA PUBLISHED REFERENCE")
        print(f"    {'':<14} {'this batch':>12} {'JNPA ref':>12}")
        print(f"    {'TAT (h)':<14} {mean_tat:>12.2f} {JNPA_ANCHOR_TAT_H:>12.2f}")
        print(f"    {'berth stay (h)':<14} {mean_stay:>12.2f} {JNPA_ANCHOR_BERTH_STAY_H:>12.2f}")
        print(f"    {'wait (h)':<14} {mean_wait:>12.2f} {JNPA_ANCHOR_WAIT_H:>12.2f}")
        print("    (a 3-row sample will not match a port-year mean; this is orientation only)")
        print()

        print("  WHY -- top drivers per vessel")
        for p in predictions:
            print(f"    row {p.row}  {p.vessel_name}   "
                  f"[{p.engine}, attribution: {p.attribution_source}]")
            base = p.breakdown.get("m3_breakdown", {}).get("base_hours")
            if base is not None:
                print(f"      {'base':<26} {float(base):>+8.2f} h")
            for d in p.drivers:
                print(f"      {str(d.get('factor')):<26} "
                      f"{float(d.get('contribution_h', 0)):>+8.2f} h   "
                      f"{d.get('coefficient')}")
            print()

    flagged = [p for p in predictions if p.flags]
    if flagged:
        print("  FLAGS")
        for p in flagged:
            print(f"    row {p.row} {p.vessel_name}: {', '.join(p.flags)}")
            for n in p.notes:
                print(f"        - {n}")
        print()


# ---------------------------------------------------------------------------
# self-test
# ---------------------------------------------------------------------------


def _self_test() -> List[Tuple[str, bool, str]]:
    import tempfile

    checks: List[Tuple[str, bool, str]] = []

    def check(name: str, ok: bool, detail: str = "") -> None:
        checks.append((name, bool(ok), detail))

    sample = jnpa_paths.SAMPLE_INPUT_XLSX
    if not os.path.exists(sample):
        check("sample_present", False, "Vessel_Training_Input_Sample.xlsx not found")
        return checks

    batch = jio.load_input(sample)
    check("input_loads", batch.ok, f"{len(batch.valid_rows)} valid rows")

    predictor, prov = resolve_predictor(None, jnpa_paths.TRAINED_MODELS_DIR)
    check("model_resolves", prov["mode"] in ("TRAINED_ARTIFACT", "UNTRAINED_FALLBACK"),
          f"{prov['mode']} / {prov['engine']}")

    preds, ctx = predict_batch(batch, predictor, wait_model="optimiser")
    check("all_rows_scored", len(preds) == len(batch.valid_rows),
          f"{len(preds)}/{len(batch.valid_rows)}")

    ok_arith = all(
        p.ata_utc is None or p.etd_utc is None
        or abs((p.etd_utc - p.ata_utc).total_seconds() / 3600.0 - p.tat_hours) < 1e-6
        for p in preds
    )
    check("etd_equals_ata_plus_tat", ok_arith, "ETD - ATA == TAT for every row")

    ok_etb = all(
        p.etb_utc is None or p.ata_utc is None
        or abs((p.etb_utc - p.ata_utc).total_seconds() / 3600.0 - p.wait_hours) < 1e-6
        for p in preds
    )
    check("etb_equals_ata_plus_wait", ok_etb, "ETB - ATA == wait for every row")

    ok_stay = all(abs((p.tat_hours - p.wait_hours) - p.berth_stay_hours) < 1e-6 for p in preds)
    check("stay_equals_tat_minus_wait", ok_stay, "berth stay reconciles on every row")

    check("ordering_ata_etb_etd",
          all(p.ata_utc <= p.etb_utc <= p.etd_utc for p in preds if p.ata_utc),
          "ATA <= ETB <= ETD")
    check("stay_above_floor",
          all(p.berth_stay_hours >= MIN_BERTH_STAY_H - 1e-9 for p in preds),
          f"every berth stay >= {MIN_BERTH_STAY_H} h")
    check("bands_ordered",
          all(p.tat_p10_hours <= p.tat_hours <= p.tat_p90_hours for p in preds),
          "p10 <= p50 <= p90")

    # Determinism: the same artefact and input must give the same answer twice.
    preds2, _ = predict_batch(batch, predictor, wait_model="optimiser")
    check("deterministic",
          all(abs(a.tat_hours - b.tat_hours) < 1e-12 for a, b in zip(preds, preds2)),
          "two runs agree to 1e-12")

    # The lower-bound flag must fire on a tiny sample -- that is the whole point.
    check("lower_bound_flag_fires",
          any("WAIT_IS_LOWER_BOUND" in p.flags for p in preds),
          "a 3-row file cannot show real berth contention, and says so")

    # queue wait model
    preds_q, ctx_q = predict_batch(batch, predictor, wait_model="queue", wait_percentile=50)
    check("queue_wait_model", all(p.wait_hours > 0 for p in preds_q),
          f"p50 wait {preds_q[0].wait_hours:.2f} h from {ctx_q['queue']['data_source']}")
    check("queue_reconciles",
          all(abs((p.tat_hours - p.wait_hours) - p.berth_stay_hours) < 1e-6 for p in preds_q),
          "arithmetic still closes after reconciliation")

    preds_n, _ = predict_batch(batch, predictor, wait_model="none")
    check("none_wait_model",
          all(p.wait_hours == 0.0 and p.etb_utc == p.ata_utc for p in preds_n),
          "ETB == ATA when wait is disabled")

    with tempfile.TemporaryDirectory() as tmp:
        csv_path = write_csv(preds, os.path.join(tmp, "p.csv"))
        with open(csv_path, encoding="utf-8") as fh:
            lines = fh.read().strip().splitlines()
        check("csv_written", len(lines) == len(preds) + 1, f"{len(lines)} lines")
        check("csv_has_targets",
              all(t in lines[0] for t in ("ETB_IST", "TAT_Hours", "ETD_IST")),
              "ETB / TAT / ETD present in the header")

        json_path = write_json(preds, os.path.join(tmp, "p.json"), batch, ctx, prov)
        with open(json_path, encoding="utf-8") as fh:
            payload = json.load(fh)
        check("json_written", len(payload["predictions"]) == len(preds),
              f"{len(payload['predictions'])} predictions")
        check("json_carries_breakdown",
              "targets" in payload["predictions"][0]["breakdown"],
              "per-row target arithmetic is in the JSON")

        xlsx_path = write_xlsx(preds, os.path.join(tmp, "p.xlsx"), ctx, prov)
        check("xlsx_written", os.path.exists(xlsx_path), os.path.basename(xlsx_path))

    return checks


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def main(argv: Optional[Sequence[str]] = None) -> int:
    p = argparse.ArgumentParser(
        description="Predict ETB / TAT / ETD for every vessel call in an input file.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            "examples:\n"
            "  python run.py predict --input data/input/Vessel_Training_Input_Sample.xlsx\n"
            "  python run.py predict --input data.xlsx --wait-model queue --wait-percentile 90\n"
            "  python run.py predict --input data.csv --out out/ --format both\n"
            "  python run.py predict --input data.xlsx "
            "--artifact trained_models/uc1_m3_tat_lightgbm_v1.2.0.pkl\n"
        ),
    )
    p.add_argument("--input", "-i", help="input .xlsx / .csv / .json")
    p.add_argument("--sheet", help="worksheet name (default: first sheet)")
    p.add_argument("--artifact", "-a", help="trained .pkl (default: newest in --model-dir)")
    p.add_argument("--model-dir", default=jnpa_paths.TRAINED_MODELS_DIR,
                   help="where to look for a trained artefact")
    p.add_argument("--wait-model", choices=("optimiser", "queue", "none"), default="optimiser",
                   help="how the pre-berth wait (and therefore ETB) is estimated")
    p.add_argument("--wait-percentile", type=int, choices=(10, 50, 90), default=50,
                   help="percentile for --wait-model queue")
    p.add_argument("--cluster-gap-hours", type=float, default=PLAN_CLUSTER_GAP_H,
                   help="requests further apart than this are planned separately")
    p.add_argument("--tide-policy", choices=("harmonic", "column", "fixed"), default="harmonic")
    p.add_argument("--tide-m", type=float, help="fixed tide height for --tide-policy fixed")
    p.add_argument("--out", "-o", default="out", help="output directory or file path")
    p.add_argument("--format", choices=("xlsx", "csv", "json", "both"), default="both",
                   help="'both' writes the spreadsheet and the full JSON")
    p.add_argument("--quiet", "-q", action="store_true", help="suppress the console report")
    p.add_argument("--selftest", action="store_true", help="run the built-in checks")
    args = p.parse_args(argv)

    if args.selftest or not args.input:
        checks = _self_test()
        passed = sum(1 for _, ok, _ in checks if ok)
        print("=" * 88)
        print(f"  predict.py self-test  |  {MODULE_VERSION}")
        print("=" * 88)
        for name, ok, detail in checks:
            print(f"  [{'PASS' if ok else 'FAIL'}] {name:<30} {detail}")
        print("-" * 88)
        print(f"  {passed}/{len(checks)} checks passed")
        if not args.input:
            print("\n  no --input given; run with --input <file> to score a data file")
        return 0 if passed == len(checks) else 1

    try:
        batch = jio.load_input(args.input, sheet=args.sheet,
                               tide_policy=args.tide_policy, fixed_tide_m=args.tide_m)
    except Exception as exc:
        print(f"ERROR reading input: {exc}", file=sys.stderr)
        return 1

    if not batch.valid_rows:
        print("ERROR: no valid rows to score. Run "
              f"'python run.py input --input {args.input} --validate' to see why.",
              file=sys.stderr)
        for issue in batch.all_issues:
            if issue.severity == "ERROR":
                print(f"  {issue}", file=sys.stderr)
        return 1

    try:
        predictor, provenance = resolve_predictor(args.artifact, args.model_dir)
    except Exception as exc:
        print(f"ERROR loading model: {exc}", file=sys.stderr)
        return 1

    predictions, context = predict_batch(
        batch, predictor,
        wait_model=args.wait_model,
        wait_percentile=args.wait_percentile,
        cluster_gap_h=args.cluster_gap_hours,
    )

    if not args.quiet:
        _print_report(predictions, batch, context, provenance)

    # Resolve --out into concrete file paths.
    out = args.out
    if os.path.splitext(out)[1]:
        stem = os.path.splitext(out)[0]
    else:
        stem = os.path.join(out, "predictions")

    written: List[str] = []
    if args.format in ("xlsx", "both"):
        written.append(write_xlsx(predictions, stem + ".xlsx", context, provenance))
    if args.format in ("csv",):
        written.append(write_csv(predictions, stem + ".csv"))
    if args.format in ("json", "both"):
        written.append(write_json(predictions, stem + ".json", batch, context, provenance))
        written.append(dashboard_path_for(stem + ".json"))

    print("  OUTPUT")
    for path in written:
        print(f"    {path}  ({os.path.getsize(path):,} bytes)")
    print()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
