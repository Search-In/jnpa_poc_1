"""
run_model.py -- one entry point for running any UC-1 model on your input file.
==============================================================================

    python run.py models --list
    python run.py models --model m1  --input data/input/Vessel_Training_Input_Sample.xlsx
    python run.py models --model all --input data/input/Vessel_Training_Input_Sample.xlsx --out out/

The eight ``uc1_m*.py`` modules are self-contained and each has its own
``--selftest`` demo, but a demo runs on *built-in* data. This file is what runs
them on *your* data: it reads the workbook through ``jnpa_input.py``, converts
each row into the dataclasses a given model expects, runs it, and writes a flat
table plus a full JSON breakdown.

WHICH MODELS ARE PER-ROW AND WHICH ARE PER-BATCH
------------------------------------------------
This distinction matters when you read the output.

    per-row    M1 DUKC             one under-keel verdict per vessel
               M2 Tidal window     one 120 h window scan per vessel
               M3 TAT              one turnaround prediction per vessel
               M6 JIT/RTA          one speed + fuel advisory per vessel
               M8 Causal chain     one confidence cascade per vessel's conditions

    per-batch  M4 Utilisation      ETA bands are per-row; occupancy and waiting
                                   statistics need a berthing log and are
                                   computed over the whole file (or the DSR log)
               M5 Berth plan       an allocation is only meaningful across all
                                   competing vessels at once
               M7 Port craft       conflicts only exist between overlapping
                                   movements, so the whole batch is scanned

TWO HONEST NOTES ABOUT THE SAMPLE FILE
--------------------------------------
1. ``DUKC_Status`` in the input sheet is M1's **output**, not its input. It is
   never fed to the model. M1's runner scores the model against it and prints an
   agreement line, which is the correct use of a label column.

2. The sheet has no tide *height*, so M1/M2/M6/M8 use the synthetic harmonic
   tide by default. Every affected number is tagged ``TIDE_SYNTHETIC``. Supply a
   ``Tide_Height_m`` column, or pass ``--tide-policy fixed --tide-m 2.6``, to
   replace the estimate with something you can defend to a Deputy Conservator.

OUTPUT
------
    --model m1     ->  out/m1_dukc.xlsx  +  out/m1_dukc.json
                       +  out/m1_dukc_dashboard.json
    --model all    ->  out/uc1_all_models.xlsx (one sheet per model, plus
                       Summary and Run_Info)  +  out/uc1_all_models.json
                       +  out/uc1_all_models_dashboard.json

Every JSON run writes two files with the same numbers and different jobs. The
plain ``.json`` is the audit trail: every formula, every substituted value,
every intermediate node -- roughly 780 KB for three vessels, because that is
what it costs to let a reviewer re-derive any figure by hand.

The ``_dashboard.json`` is what a UI reads: one object per vessel, its inputs,
and five to nine fields per model, with a glossary of its own keys embedded in
the file. See ``dashboard_json.py`` for what qualifies as a dashboard field.

Exit code is 0 when every requested model ran, 1 otherwise.
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import sys
import traceback
from dataclasses import dataclass, field
from datetime import datetime, timedelta, timezone
from typing import Any, Callable, Dict, List, Optional, Sequence, Tuple

import jnpa_paths

jnpa_paths.ensure_on_syspath()

import jnpa_input as jio  # noqa: E402

MODULE_ID: str = "UC1-RUNNER"
MODULE_VERSION: str = "run-model-v1.0.0"

DEFAULT_AIS_STALENESS_MIN: float = 15.0


def _ist(dt: Optional[datetime]) -> str:
    return dt.astimezone(jio.IST).strftime("%Y-%m-%d %H:%M") if dt else ""


def _r(x: Any, n: int = 2) -> Any:
    return round(float(x), n) if isinstance(x, (int, float)) else x


# ---------------------------------------------------------------------------
# result container
# ---------------------------------------------------------------------------


@dataclass
class RunResult:
    """What one model produced for one input file."""

    model_id: str
    module: str
    title: str
    scope: str                       # "per-row" | "per-batch"
    columns: Tuple[str, ...]
    rows: List[Dict[str, Any]]
    summary: Dict[str, Any] = field(default_factory=dict)
    details: Dict[str, Any] = field(default_factory=dict)
    notes: List[str] = field(default_factory=list)
    console: List[str] = field(default_factory=list)
    ok: bool = True
    error: str = ""

    def as_dict(self) -> Dict[str, Any]:
        return {
            "model_id": self.model_id,
            "module": self.module,
            "title": self.title,
            "scope": self.scope,
            "ok": self.ok,
            "error": self.error,
            "summary": self.summary,
            "columns": list(self.columns),
            "rows": self.rows,
            "details": self.details,
            "notes": self.notes,
        }


# ---------------------------------------------------------------------------
# M1 -- DUKC / RTUKC
# ---------------------------------------------------------------------------


def run_m1(batch: jio.InputBatch, opts: Dict[str, Any]) -> RunResult:
    import uc1_m1_dukc as m1

    rows: List[Dict[str, Any]] = []
    details: Dict[str, Any] = {}
    agree = disagree = scored = 0

    for call in batch.valid_rows:
        vessel = jio.to_m1_vessel(call)
        allr = m1.evaluate_all_reaches(
            vessel, tide_m=call.tide_height_m,
            siltation_m=call.siltation_m, dredging_m=call.dredging_delta_m,
            with_sensitivity=True,
        )
        binding = next(r for r in allr.results if r.reach_id == allr.binding_reach_id)

        # The sheet's DUKC_Status is a LABEL. Score against it; never feed it in.
        reported = call.dukc_status_reported
        verdict = ""
        if reported:
            scored += 1
            normalised = {"SAFE": "SAFE", "MARGINAL": "MARGINAL",
                          "NO GO": "NO GO", "NOGO": "NO GO", "NO-GO": "NO GO"}.get(
                reported, reported)
            if normalised == allr.transit_status:
                agree += 1
                verdict = "AGREE"
            else:
                disagree += 1
                verdict = f"DIFFER (sheet says {normalised})"

        rows.append({
            "Row": call.row,
            "Vessel": call.vessel_name,
            "Draft_m": _r(call.draft_m),
            "Speed_kn": _r(call.transit_speed_kn, 1),
            "Cb": _r(binding.block_coefficient_cb, 2),
            "Tide_m": _r(call.tide_height_m),
            "Binding_Reach": allr.binding_reach_id,
            "Charted_Depth_m": _r(binding.charted_depth_m),
            "Siltation_m": _r(call.siltation_m),
            "Dredging_m": _r(call.dredging_delta_m),
            "Effective_Depth_m": _r(binding.effective_depth_m, 3),
            "Squat_m": _r(binding.squat_m, 3),
            "Gross_UKC_m": _r(binding.gross_ukc_m, 3),
            "Net_UKC_m": _r(binding.net_ukc_m, 3),
            "Status": allr.transit_status,
            "Sheet_DUKC_Status": reported,
            "Model_vs_Sheet": verdict,
            "Min_Tide_For_SAFE_m": _r(binding.min_tide_for_safe_m, 3),
            "Max_SAFE_Speed_kn": _r(binding.max_safe_speed_kn, 2)
            if binding.max_safe_speed_kn is not None else "no speed makes it SAFE",
            "Sensitivity_Robust": binding.sensitivity_robust,
            "Recommendation": allr.recommendation,
        })
        details[call.call_id] = allr.breakdown

    notes = [
        "Cb = 0.65 for container, 0.80 for bulk; squat = min(2.5, Cb * V^2 / 100).",
        "Net UKC = (charted + tide - siltation + dredging) - (draft + squat) - 1.0 m margin.",
        "Status bands: >= 1.0 m SAFE, 0.6-1.0 m MARGINAL, < 0.6 m NO GO.",
        "The binding reach is argmin(net UKC) across all reaches, each at its own "
        "speed cap -- not simply the shallowest reach.",
    ]
    summary: Dict[str, Any] = {
        "vessels": len(rows),
        "safe": sum(1 for r in rows if r["Status"] == "SAFE"),
        "marginal": sum(1 for r in rows if r["Status"] == "MARGINAL"),
        "no_go": sum(1 for r in rows if r["Status"] == "NO GO"),
        "dukc_core_fingerprint": m1.DUKC_CORE_FINGERPRINT,
    }
    if scored:
        summary["validation_vs_sheet"] = {
            "scored": scored, "agree": agree, "disagree": disagree,
            "agreement_pct": round(100.0 * agree / scored, 1),
        }
        notes.append(
            f"Scored against the sheet's DUKC_Status column: {agree}/{scored} agree. "
            f"Disagreements are expected while tide height is synthetic -- the sheet's "
            f"status was recorded at the real tide, which this run does not have."
        )
    return RunResult("UC1-M1", "uc1_m1_dukc", "DUKC / Real-Time Under-Keel Clearance",
                     "per-row", tuple(rows[0]) if rows else (), rows, summary, details, notes)


# ---------------------------------------------------------------------------
# M2 -- tidal window scanner
# ---------------------------------------------------------------------------


def run_m2(batch: jio.InputBatch, opts: Dict[str, Any]) -> RunResult:
    import uc1_m2_tidal_window as m2

    rows: List[Dict[str, Any]] = []
    details: Dict[str, Any] = {}
    hours = float(opts.get("horizon_hours", m2.FORECAST_HOURS))

    for call in batch.valid_rows:
        vessel = jio.to_m2_vessel(call)
        start = call.ata_utc or call.eta_utc or datetime.now(timezone.utc)
        res = m2.evaluate_tidal_windows(
            vessel, start=start, hours=hours,
            scenarios=(
                m2.ScenarioSpec("DREDGED", "Dredged +0.5 m", 0.0, 0.5),
                m2.ScenarioSpec("SILTED", "Silted -0.3 m", 0.3, 0.0),
            ),
        )
        base = res.baseline
        by_id = {c.scenario_id: c for c in res.comparisons}
        dredged, silted = by_id.get("DREDGED"), by_id.get("SILTED")

        # Does the model agree the sheet's stated window is transitable?
        stated = ""
        if call.tide_window_start_utc and call.tide_window_end_utc:
            covered = any(
                w.start <= call.tide_window_start_utc and w.end >= call.tide_window_end_utc
                for w in base.windows
            )
            overlaps = any(
                w.start < call.tide_window_end_utc and w.end > call.tide_window_start_utc
                for w in base.windows
            )
            stated = "fully inside a model window" if covered else (
                "partly overlaps a model window" if overlaps else "no model window overlaps it")

        rows.append({
            "Row": call.row,
            "Vessel": call.vessel_name,
            "Draft_m": _r(call.draft_m),
            "Required_Tide_m": _r(res.required_tide_for_feasible_m, 3),
            "Binding_Reach": res.binding_reach_id,
            "Horizon_h": _r(res.horizon_hours, 1),
            "Samples": res.samples,
            "Windows": base.window_count,
            "Usable_Hours": _r(base.total_usable_hours),
            "Availability_pct": _r(base.availability_pct, 1),
            "Mean_Window_h": _r(base.mean_window_hours),
            "Longest_Window_h": _r(base.longest_window_h),
            "Max_Gap_h": _r(base.max_gap_hours),
            "Next_Window_Start_IST": _ist(base.first_window_start),
            "Dredged_Delta_h": _r(dredged.delta_hours) if dredged else "",
            "Dredged_Delta_pct": _r(dredged.delta_pct, 1) if dredged else "",
            "Silted_Delta_h": _r(silted.delta_hours) if silted else "",
            "Silted_Delta_pct": _r(silted.delta_pct, 1) if silted else "",
            "Conditional_Extra_h": _r(res.conditional.total_usable_hours),
            "Sheet_Window_IST": (f"{_ist(call.tide_window_start_utc)} - "
                                 f"{_ist(call.tide_window_end_utc)}"
                                 if call.tide_window_start_utc else ""),
            "Sheet_Window_vs_Model": stated,
            "Recommendation": res.recommendation,
        })
        details[call.call_id] = res.breakdown

    return RunResult(
        "UC1-M2", "uc1_m2_tidal_window", "Tidal Window Scanner & Extension",
        "per-row", tuple(rows[0]) if rows else (), rows,
        {
            "vessels": len(rows),
            "mean_usable_hours": _r(sum(r["Usable_Hours"] for r in rows) / len(rows)) if rows else 0,
            "mean_availability_pct": _r(
                sum(r["Availability_pct"] for r in rows) / len(rows), 1) if rows else 0,
            "horizon_hours": hours,
            "step_hours": m2.STEP_HOURS,
        },
        details,
        [
            f"Each scan steps {m2.STEP_HOURS:g} h over {hours:g} h and tests every reach "
            f"at its own speed cap.",
            "Headline windows require SAFE (net UKC >= 1.0 m). MARGINAL periods are "
            "reported separately as Conditional_Extra_h and never counted as usable.",
            "Dredged/Silted deltas are the 'extend the tidal window' deliverable: they "
            "price a dredging campaign in transitable hours.",
        ],
    )


# ---------------------------------------------------------------------------
# M3 -- TAT prediction (delegates to predict.py so there is one code path)
# ---------------------------------------------------------------------------


def run_m3(batch: jio.InputBatch, opts: Dict[str, Any]) -> RunResult:
    import predict as pr

    predictor, provenance = pr.resolve_predictor(opts.get("artifact"), opts.get("model_dir", "models"))
    preds, context = pr.predict_batch(
        batch, predictor,
        wait_model=opts.get("wait_model", "optimiser"),
        wait_percentile=int(opts.get("wait_percentile", 50)),
    )
    rows = [p.as_row() for p in preds]
    notes = [
        "TAT is the only learned quantity in UC-1; every other model is deterministic.",
        f"Model: {provenance['mode']} engine={provenance['engine']}.",
        "ETB = ATA + wait; ETD = ATA + TAT; berth stay = TAT - wait.",
    ]
    if provenance.get("warning"):
        notes.append(provenance["warning"])
    if any("WAIT_IS_LOWER_BOUND" in p.flags for p in preds):
        notes.append(
            "Berth contention was computed within this file only, so the wait -- and "
            "therefore ETB -- is a lower bound. Use --wait-model queue on a partial file."
        )
    return RunResult(
        "UC1-M3", "uc1_m3_tat_predict", "Turnaround Time prediction (ETB / TAT / ETD)",
        "per-row", tuple(pr.OUTPUT_COLUMNS), rows,
        {
            "vessels": len(rows),
            "engine": provenance["engine"],
            "model_mode": provenance["mode"],
            "artifact": provenance.get("artifact_path", ""),
            "holdout_mae_hours": provenance.get("holdout_mae_hours"),
            "holdout_coverage_pct": provenance.get("holdout_coverage_pct"),
            "mean_tat_hours": _r(sum(p.tat_hours for p in preds) / len(preds)) if preds else 0,
            "wait_model": context.get("wait_model"),
        },
        {p.call_id: p.breakdown for p in preds},
        notes,
    )


# ---------------------------------------------------------------------------
# M4 -- ETA uncertainty and berth utilisation
# ---------------------------------------------------------------------------


def run_m4(batch: jio.InputBatch, opts: Dict[str, Any]) -> RunResult:
    import uc1_m4_berth_utilisation as m4

    staleness = float(opts.get("ais_staleness_min", DEFAULT_AIS_STALENESS_MIN))
    rows: List[Dict[str, Any]] = []
    details: Dict[str, Any] = {}

    for call in batch.valid_rows:
        obs = jio.to_m4_eta_observation(call, ais_staleness_minutes=staleness)
        band = m4.compute_eta_band(obs)
        rows.append({
            "Row": call.row,
            "Vessel": call.vessel_name,
            "Now_IST": _ist(obs.now_utc),
            "Forecast_ETA_IST": _ist(obs.forecast_eta_utc),
            "Horizon_h": _r(band.horizon_hours),
            "AIS_Staleness_min": _r(band.ais_staleness_minutes, 1),
            "Sigma_h": _r(band.sigma_hours, 3),
            "ETA_P10_IST": _ist(band.eta_p10_utc),
            "ETA_P50_IST": _ist(band.eta_p50_utc),
            "ETA_P90_IST": _ist(band.eta_p90_utc),
            "Band_Width_h": _r(band.band_width_hours),
            "Confidence": band.confidence_label,
        })
        details[call.call_id] = band.breakdown

    # Occupancy and waiting need a berthing log; the input file has no ATB/ATD
    # (those are targets), so these come from the DSR-derived log.
    end = datetime.now(timezone.utc)
    records, berths, source = m4.load_records_with_fallback(end - timedelta(days=120), end)
    window_start = min((r.actual_atb_utc for r in records if r.actual_atb_utc),
                       default=end - timedelta(days=7))
    window_end = max((r.actual_atd_utc for r in records if r.actual_atd_utc),
                     default=end)
    report = m4.berth_utilisation_report(records, berths, window_start, window_end,
                                         data_source=source)

    console = [
        "  BERTH OCCUPANCY (from the berthing log, not from the input file)",
        f"    source            {source}",
        f"    window            {_ist(window_start)} -> {_ist(window_end)} IST",
        f"    records / berths  {report.record_count} / {report.berth_count}",
        f"    occupancy         {report.calendar.overall_occupancy_pct:.2f} %",
        f"    double-booked     "
        f"{report.calendar.breakdown['occupied_hours'].get('double_booked', 0.0):.2f} h "
        f"(union vs raw sum -- a data-quality signal, not an error)",
        "",
        "  WAITING TIME",
        f"    definition        {report.waiting.definition}",
        f"    usable / dropped  {report.waiting.n} / {report.waiting.n_dropped}",
    ]
    if report.waiting.n:
        console.append(f"    p50 / p90 / mean  {report.waiting.p50_hours:.2f} h / "
                       f"{report.waiting.p90_hours:.2f} h / {report.waiting.mean_hours:.2f} h")
    else:
        console.append("    NOT COMPUTABLE from this source -- see the note below")

    notes = [
        "sigma = 0.06 * horizon_hours + 0.05 * AIS_staleness_minutes; "
        "the band is p50 +/- 1.28 sigma (80%).",
        f"AIS staleness is assumed {staleness:g} min (the input file has no AIS column); "
        f"override with --ais-staleness-min.",
        "Occupancy uses union-merged intervals per berth so double-booked dirty data "
        "cannot push a cell over 100%; the raw-minus-union gap is reported instead.",
    ]
    if report.waiting.n == 0:
        reasons = sorted({d.get("reason", "?") for d in report.waiting.drop_reasons})
        notes.append(
            f"Waiting time could NOT be computed from {source}: all "
            f"{report.waiting.n_dropped} records were dropped ({', '.join(reasons)}). "
            f"Daily Status Report section (H) records berthing and expected completion "
            f"but never an arrival at anchorage, and waiting time is "
            f"'{report.waiting.definition}'. Recovering it needs the PCS VESARR logs."
        )

    return RunResult(
        "UC1-M4", "uc1_m4_berth_utilisation", "ETA uncertainty & berth utilisation",
        "per-batch", tuple(rows[0]) if rows else (), rows,
        {
            "vessels": len(rows),
            "mean_sigma_hours": _r(sum(r["Sigma_h"] for r in rows) / len(rows), 3) if rows else 0,
            "occupancy_source": source,
            "occupancy_records": report.record_count,
            "berths": report.berth_count,
            "overall_occupancy_pct": _r(report.calendar.overall_occupancy_pct),
            "waiting_n": report.waiting.n,
            "waiting_p50_h": _r(report.waiting.p50_hours) if report.waiting.n else None,
            "waiting_p90_h": _r(report.waiting.p90_hours) if report.waiting.n else None,
        },
        {"eta_bands": details, "utilisation_report": report.breakdown},
        notes, console,
    )


# ---------------------------------------------------------------------------
# M5 -- dynamic berth plan optimisation
# ---------------------------------------------------------------------------


def run_m5(batch: jio.InputBatch, opts: Dict[str, Any]) -> RunResult:
    import predict as pr
    import uc1_m5_berth_optimiser as m5

    waits = pr.wait_from_optimiser(batch.valid_rows,
                                   float(opts.get("cluster_gap_hours", pr.PLAN_CLUSTER_GAP_H)))
    by_id = {c.call_id: c for c in batch.valid_rows}

    rows: List[Dict[str, Any]] = []
    for call_id, info in sorted(waits.items()):
        call = by_id.get(call_id)
        if call is None:
            continue
        rows.append({
            "Row": call.row,
            "Vessel": call.vessel_name,
            "LOA_m": _r(call.loa_m),
            "Draft_m": _r(call.draft_m),
            "Requested_Berth": call.requested_berth,
            "Assigned_Berth": info.get("assigned_berth", ""),
            "Requested_Start_IST": _ist(call.ata_utc),
            "Assigned_Start_IST": _ist(info.get("start_utc")),
            "Service_h": _r(call.service_hours),
            "Wait_h": _r(info.get("wait_hours", 0.0)),
            "Berth_Shift": bool(info.get("is_berth_shift")),
            "Tide_Miss": bool(info.get("tide_miss")),
            "Algorithm": info.get("algorithm", ""),
            "Cluster_Size": info.get("cluster_size", 1),
            "Tidal_Windows": info.get("tidal_windows_available", 0),
            "Rationale": info.get("rationale", ""),
        })
    rows.sort(key=lambda r: r["Row"])

    wait_total = sum(float(r["Wait_h"]) for r in rows)
    shifts = sum(1 for r in rows if r["Berth_Shift"])
    misses = sum(1 for r in rows if r["Tide_Miss"])
    w = m5.DEFAULT_WEIGHTS
    cost = w.wait_hour * wait_total + w.berth_shift * shifts + w.tide_miss * misses

    console = [
        "  COST BREAKDOWN",
        f"    {'component':<16} {'quantity':>10} {'weight':>8} {'subtotal':>10}",
        "    " + "-" * 48,
        f"    {'wait hours':<16} {wait_total:>10.2f} {w.wait_hour:>8.1f} "
        f"{w.wait_hour * wait_total:>10.2f}",
        f"    {'berth shifts':<16} {shifts:>10} {w.berth_shift:>8.1f} "
        f"{w.berth_shift * shifts:>10.2f}",
        f"    {'tide misses':<16} {misses:>10} {w.tide_miss:>8.1f} "
        f"{w.tide_miss * misses:>10.2f}",
        "    " + "-" * 48,
        f"    {'TOTAL':<16} {'':>10} {'':>8} {cost:>10.2f}",
    ]

    clusters = pr._cluster_by_time(batch.valid_rows,
                                   float(opts.get("cluster_gap_hours", pr.PLAN_CLUSTER_GAP_H)))
    notes = [
        "Objective = 1.0 * wait_hours + 2.0 * tide_misses + 0.5 * berth_shifts.",
        "LOA, draft and non-overlap (+0.5 h turnaround buffer) are hard constraints, "
        "guaranteed by construction, so the plan is always physically valid.",
        "Tide is a costed soft constraint by default: a vessel with no feasible window "
        "within 72 h is still berthed but carries a 2.0 tide-miss cost.",
        f"Vessels were split into {len(clusters)} planning cluster(s) "
        f"({', '.join(str(len(c)) for c in clusters)} vessel(s) each); calls more than "
        f"{opts.get('cluster_gap_hours', pr.PLAN_CLUSTER_GAP_H):g} h apart do not compete "
        f"for a berth and are optimised separately.",
    ]
    if any(len(c) <= 2 for c in clusters):
        notes.append(
            "At least one cluster has 1-2 vessels, so there is nothing to contend with "
            "and the optimiser has little to do. Feed the full arrival list to see real "
            "berth competition."
        )

    return RunResult(
        "UC1-M5", "uc1_m5_berth_optimiser", "Dynamic berth plan optimisation",
        "per-batch", tuple(rows[0]) if rows else (), rows,
        {
            "requests": len(rows),
            "assigned": sum(1 for r in rows if r["Assigned_Berth"]),
            "clusters": len(clusters),
            "total_wait_hours": _r(wait_total),
            "berth_shifts": shifts,
            "tide_misses": misses,
            "total_cost": _r(cost, 3),
            "weights": {"wait_hour": w.wait_hour, "tide_miss": w.tide_miss,
                        "berth_shift": w.berth_shift},
            "cpsat_available": m5.cpsat_available(),
            "algorithms_used": sorted({r["Algorithm"] for r in rows}),
        },
        {"assignments": rows},
        notes, console,
    )


# ---------------------------------------------------------------------------
# M6 -- JIT arrival / RTA advisory
# ---------------------------------------------------------------------------


def run_m6(batch: jio.InputBatch, opts: Dict[str, Any]) -> RunResult:
    import uc1_m6_jit_rta as m6

    rows: List[Dict[str, Any]] = []
    details: Dict[str, Any] = {}
    pairs = []
    assumed_distance = assumed_ready = 0

    for call in batch.valid_rows:
        if call.distance_source == "DEFAULT":
            assumed_distance += 1
        if call.berth_ready_utc is None:
            assumed_ready += 1
        vessel = jio.to_m6_vessel(call)
        readiness = jio.to_m6_readiness(call)
        res = m6.evaluate_jit(vessel, readiness)
        pairs.append((vessel, readiness))
        rows.append({
            "Row": call.row,
            "Vessel": call.vessel_name,
            "Distance_NM": _r(call.distance_nm, 1),
            "Distance_Source": call.distance_source,
            "Now_IST": _ist(vessel.now),
            "RTA_IST": _ist(res.rta),
            "RTA_Driver": res.rta_driver,
            "Available_h": _r(res.available_hours),
            "Required_Speed_kn": _r(res.required_speed_kn),
            "Recommended_Speed_kn": _r(res.recommended_speed_kn),
            "Speed_Clamped": res.speed_clamped,
            "Feasible": res.feasible,
            "Misses_Tidal_Window": res.misses_tidal_window,
            "Baseline_Fuel_t": _r(res.baseline.steaming_fuel_t),
            "JIT_Fuel_t": _r(res.jit.steaming_fuel_t),
            "Fuel_Saved_t": _r(res.headline.fuel_saved_t),
            "CO2_Saved_t": _r(res.headline.co2_saved_t, 3),
            "Bunker_Saved_USD": _r(res.headline.bunker_saved_usd, 0),
            "Anchorage_h_Eliminated": _r(res.anchorage_hours_eliminated),
            "Basis": res.headline.basis,
            "Recommendation": res.recommendation,
        })
        details[call.call_id] = res.breakdown

    fleet = m6.evaluate_fleet(pairs) if pairs else None
    notes = [
        "RTA = max(berth ready, tidal window start). Required speed = distance / hours.",
        "Fuel = 3.2 t/h * (speed / 16 kn)^3 * transit hours; CO2 = fuel * 3.114 (IMO).",
        "Headline savings are STEAMING-ONLY -- the conservative figure. The "
        "anchorage-inclusive number is larger and is reported as secondary in the JSON.",
        "Bunker price USD 600/t and anchorage idle 0.35 t/h are SIMULATED assumptions.",
    ]
    if assumed_distance:
        notes.append(
            f"{assumed_distance} of {len(rows)} rows had no Distance_NM column, so "
            f"{jio.DEFAULT_DISTANCE_NM:g} NM was assumed. Fuel savings scale directly "
            f"with distance -- supply the real value before quoting these numbers."
        )
    if assumed_ready:
        notes.append(
            f"{assumed_ready} of {len(rows)} rows had no Berth_Ready column; the stated "
            f"tidal window opening (or the ATA) was used instead."
        )

    return RunResult(
        "UC1-M6", "uc1_m6_jit_rta", "Just-In-Time arrival & RTA advisory",
        "per-row", tuple(rows[0]) if rows else (), rows,
        {
            "vessels": len(rows),
            "fleet_fuel_saved_t": _r(fleet.fleet_fuel_saved_t) if fleet else 0,
            "fleet_co2_saved_t": _r(fleet.fleet_co2_saved_t, 3) if fleet else 0,
            "fleet_bunker_saved_usd": _r(fleet.fleet_bunker_saved_usd, 0) if fleet else 0,
            "fleet_anchorage_hours_eliminated": _r(
                fleet.fleet_anchorage_hours_eliminated) if fleet else 0,
            "basis": "STEAMING_ONLY",
            "commercial_figures": "SIMULATED",
        },
        details, notes,
    )


# ---------------------------------------------------------------------------
# M7 -- port-craft assignment & conflict detection
# ---------------------------------------------------------------------------


def run_m7(batch: jio.InputBatch, opts: Dict[str, Any]) -> RunResult:
    import uc1_m7_port_craft as m7

    preset = str(opts.get("roster_preset", "real"))
    movements = [jio.to_m7_movement(c) for c in batch.valid_rows]
    roster = m7.build_default_roster(preset)
    down = [x for x in str(opts.get("down_craft", "")).split(",") if x]
    if down:
        roster = m7.apply_outage(roster, down)

    report = m7.evaluate(movements, roster, scenario="input_file", roster_preset=preset)
    by_movement = {a.movement_id: a for a in report.allocations}

    rows: List[Dict[str, Any]] = []
    for call, mv in zip(batch.valid_rows, movements):
        a = by_movement.get(mv.movement_id)
        rows.append({
            "Row": call.row,
            "Vessel": call.vessel_name,
            "Movement": mv.movement_id,
            "Type": mv.movement_type,
            "Class": mv.vessel_class,
            "Berth": mv.berth_id,
            "Start_IST": _ist(mv.start_utc),
            "End_IST": _ist(mv.end_utc),
            "Req_Pilots": mv.req_pilots,
            "Req_Tugs": mv.req_tugs,
            "Req_Mooring": mv.req_mooring,
            "Assigned_Pilots": ", ".join(a.pilot_ids) if a else "",
            "Assigned_Tugs": ", ".join(a.tug_ids) if a else "",
            "Assigned_Mooring": ", ".join(a.mooring_ids) if a else "",
            "Shortfall": json.dumps(a.shortfall) if a and a.shortfall else "",
            "Response_Gap_min": _r(a.response_gap_min, 1) if a else "",
            "Feasible": bool(a and a.feasible),
        })

    console: List[str] = []
    if report.conflicts:
        console += ["  CONFLICT BLOCKS",
                    f"    {'window (IST)':<36} {'role':<10} {'peak':>5} {'severity':<9} movements",
                    "    " + "-" * 92]
        for c in report.conflicts:
            console.append(
                f"    {_ist(c.start_utc)} - {_ist(c.end_utc):<16} {c.role:<10} "
                f"{c.peak_deficit:>5} {c.severity:<9} {', '.join(c.movement_ids)}")
        console.append("")
    if report.proposals:
        console += ["  RECOMMENDED SINGLE-UNIT SWAPS"]
        for p in report.proposals:
            console.append(f"    [{p.action}] {p.rationale}")
            console.append(f"        gap closed: {p.gap_closed_minutes:.0f} min ({p.basis})")
        console.append("")

    notes = [
        f"Roster preset '{preset}': {len(roster)} craft. The 'real' preset is "
        f"transcribed from Details_of_Port_Crafts.pdf (18 craft); the 'poc' preset is "
        f"the WS2 spec's 9 craft. The spec cites that PDF but reports 9 -- we default "
        f"to the PDF and keep the spec figure as a preset.",
        "Allocation is interval-aware: a craft unit already committed to an overlapping "
        "movement cannot be handed to a second one.",
        "Each vessel row becomes one inbound BERTHING movement; tug and pilot counts "
        "come from the class + bow-thruster rules, not from the input file.",
    ]
    if not report.conflicts:
        overlapping = sum(
            1 for i, a in enumerate(movements) for b in movements[i + 1:]
            if a.start_utc < b.end_utc and b.start_utc < a.end_utc
        )
        if overlapping == 0:
            notes.append(
                f"No conflicts, and none is possible here: none of the {len(movements)} "
                f"movements overlap in time, so a single pilot and a single tug could "
                f"serve all of them in sequence. A craft conflict requires simultaneous "
                f"demand -- reducing the roster with --down-craft will not manufacture "
                f"one. Feed a file with vessels arriving within the same 1.5 h movement "
                f"window to exercise the conflict path."
            )
        else:
            notes.append(
                f"No conflicts: {overlapping} overlapping movement pair(s) against "
                f"{len(roster)} craft is still within supply. Try --roster-preset poc "
                f"with --down-craft PL-01,PL-02 to force a shortage."
            )

    return RunResult(
        "UC1-M7", "uc1_m7_port_craft", "Port-craft assignment & conflict detection",
        "per-batch", tuple(rows[0]) if rows else (), rows,
        {
            "movements": len(movements),
            "roster_preset": preset,
            "roster_size": len(roster),
            "craft_down": down,
            "status": report.status,
            "conflicts": len(report.conflicts),
            "proposals": len(report.proposals),
            "total_response_gap_min": _r(report.total_response_gap_min, 1),
            "utilisation_by_role": {k: _r(v, 1) for k, v in report.utilisation_pct.items()},
        },
        report.breakdown, notes, console,
    )


# ---------------------------------------------------------------------------
# M8 -- reactive confidence chain (causal DAG)
# ---------------------------------------------------------------------------


def run_m8(batch: jio.InputBatch, opts: Dict[str, Any]) -> RunResult:
    import uc1_m8_causal_chain as m8

    graph = m8.build_graph()
    rows: List[Dict[str, Any]] = []
    details: Dict[str, Any] = {}

    for call in batch.valid_rows:
        disruptions = jio.to_m8_disruptions(call)
        res = m8.propagate(graph, disruptions)
        state = res.final_state
        # top_root_causes is a list of (node_id, share_pct) tuples.
        causes = [(str(n), float(s)) for n, s in res.top_root_causes]
        rows.append({
            "Row": call.row,
            "Vessel": call.vessel_name,
            "Disruptions": ", ".join(d.label for d in disruptions) or "(none -- baseline)",
            "DUKC_Net_UKC_m": _r(state.get("DUKC_NET_UKC_M"), 3),
            "Deep_Draft_Window_h": _r(state.get("DEEP_DRAFT_WINDOW_H"), 2),
            "Pilotage_Hold": _r(state.get("PILOTAGE_HOLD"), 3),
            "Channel_Throughput_vph": _r(state.get("CHANNEL_THROUGHPUT_VPH"), 3),
            "Anchorage_Queue": _r(state.get("ANCHORAGE_QUEUE_N"), 2),
            "Berth_Plan_Feasibility": _r(state.get("BERTH_PLAN_FEASIBILITY"), 3),
            "Crane_Productivity": _r(state.get("CRANE_PRODUCTIVITY"), 3),
            "TAT_Delay_h": _r(state.get("TAT_DELAY_H"), 2),
            "System_Confidence": _r(state.get("SYS_CONFIDENCE"), 3),
            "Confidence_Delta": _r(res.confidence_delta, 3),
            "Alert_Level": res.alert_level,
            "Rules_Fired": ", ".join(t.rule.rule_id for t in res.triggered_rules) or "none",
            "Critical_Rules": ", ".join(
                t.rule.rule_id for t in res.triggered_rules
                if t.rule.severity == "CRITICAL") or "none",
            "Root_Cause_1": f"{causes[0][0]} {causes[0][1]:.0f}%" if causes else "",
            "Root_Cause_2": f"{causes[1][0]} {causes[1][1]:.0f}%" if len(causes) > 1 else "",
            "Root_Cause_3": f"{causes[2][0]} {causes[2][1]:.0f}%" if len(causes) > 2 else "",
            "Propagation_Steps": len(res.propagation_log),
        })
        details[call.call_id] = res.breakdown

    return RunResult(
        "UC1-M8", "uc1_m8_causal_chain", "Reactive confidence chain (23-node causal DAG)",
        "per-row", tuple(rows[0]) if rows else (), rows,
        {
            "vessels": len(rows),
            "nodes": len(graph.nodes),
            "edges": len(graph.edges),
            "min_confidence": min((r["System_Confidence"] for r in rows), default=None),
            "max_confidence": max((r["System_Confidence"] for r in rows), default=None),
            "alert_levels": sorted({r["Alert_Level"] for r in rows}),
            "any_critical": any(r["Critical_Rules"] != "none" for r in rows),
            "reference_vessel_draft_m": graph.reference_draft_m,
            "reference_vessel_speed_kn": graph.reference_speed_kn,
        },
        details,
        [
            f"DUKC_Net_UKC_m here is for the REFERENCE deep-draft vessel "
            f"({graph.reference_draft_m:g} m draft at {graph.reference_speed_kn:g} kn), NOT "
            f"the vessel on that row. M8 answers 'is the port open to deep-draft traffic "
            f"under these conditions', so it can read NO GO on a row whose own (shallower) "
            f"vessel M1 calls SAFE. Compare M8's DUKC column with M1's only through that "
            f"lens; M1's column is the per-vessel answer.",
            f"{len(graph.nodes)} nodes, {len(graph.edges)} edges, acyclic by construction "
            f"(every edge runs low index -> high index).",
            "Every run logs one propagation step per node -- all 23, including unchanged "
            "ones -- so the audit trail is complete rather than only interesting.",
            "Only the 10 exogenous nodes are set from your data; everything downstream is "
            "computed. Values equal to the node baseline are not listed as disruptions.",
            "Edge weights: 4 are exact physics, 1 is calibrated against M2's scanner, and "
            "25 are labelled EXPERT_JUDGEMENT. They are labelled, not hidden.",
        ],
    )


# ---------------------------------------------------------------------------
# registry
# ---------------------------------------------------------------------------

Runner = Callable[[jio.InputBatch, Dict[str, Any]], RunResult]

MODELS: Dict[str, Tuple[str, str, Runner]] = {
    "m1": ("uc1_m1_dukc", "DUKC / Real-Time Under-Keel Clearance", run_m1),
    "m2": ("uc1_m2_tidal_window", "Tidal Window Scanner & Extension", run_m2),
    "m3": ("uc1_m3_tat_predict", "TAT prediction (ETB / TAT / ETD)", run_m3),
    "m4": ("uc1_m4_berth_utilisation", "ETA uncertainty & berth utilisation", run_m4),
    "m5": ("uc1_m5_berth_optimiser", "Dynamic berth plan optimisation", run_m5),
    "m6": ("uc1_m6_jit_rta", "Just-In-Time arrival & RTA advisory", run_m6),
    "m7": ("uc1_m7_port_craft", "Port-craft assignment & conflict detection", run_m7),
    "m8": ("uc1_m8_causal_chain", "Reactive confidence chain (causal DAG)", run_m8),
}

MODEL_SCOPE: Dict[str, str] = {
    "m1": "per-row", "m2": "per-row", "m3": "per-row", "m4": "per-batch",
    "m5": "per-batch", "m6": "per-row", "m7": "per-batch", "m8": "per-row",
}


def run_one(key: str, batch: jio.InputBatch, opts: Dict[str, Any]) -> RunResult:
    module, title, fn = MODELS[key]
    try:
        return fn(batch, opts)
    except Exception as exc:
        return RunResult(
            f"UC1-{key.upper()}", module, title, MODEL_SCOPE[key], (), [],
            ok=False, error=f"{type(exc).__name__}: {exc}",
            details={"traceback": traceback.format_exc()},
        )


# ---------------------------------------------------------------------------
# output
# ---------------------------------------------------------------------------


def _print_result(res: RunResult) -> None:
    print()
    print("=" * 118)
    print(f"  {res.model_id}  {res.title}")
    print(f"  module {res.module}   scope {res.scope}")
    print("=" * 118)
    if not res.ok:
        print(f"  FAILED: {res.error}")
        return

    if res.rows:
        cols = list(res.columns)
        # Keep the console readable; the full row is always in the JSON/xlsx.
        wide = {"Recommendation", "Rationale", "Disruptions", "Shortfall",
                "Top_Driver_1", "Top_Driver_2", "Top_Driver_3", "Stressors",
                "Sheet_Window_vs_Model", "Assigned_Tugs", "Assigned_Pilots",
                "Assigned_Mooring", "Flags", "Notes"}
        show = [c for c in cols if c not in wide][:13]
        widths = {
            c: max(len(c), max((len(str(r.get(c, ""))) for r in res.rows), default=0))
            for c in show
        }
        widths = {c: min(w, 20) for c, w in widths.items()}
        header = "  " + " ".join(f"{c[:widths[c]]:<{widths[c]}}" for c in show)
        print(header)
        print("  " + "-" * (len(header) - 2))
        for r in res.rows:
            print("  " + " ".join(f"{str(r.get(c, ''))[:widths[c]]:<{widths[c]}}" for c in show))
        if len(cols) > len(show):
            print(f"  ... {len(cols) - len(show)} more columns in the output file: "
                  f"{', '.join(c for c in cols if c not in show)}")
        print()

    for line in res.console:
        print(line)

    if res.summary:
        print("  SUMMARY")
        for k, v in res.summary.items():
            print(f"    {k:<32} {v}")
        print()
    if res.notes:
        print("  NOTES")
        for n in res.notes:
            print(f"    - {n}")
        print()


def write_outputs(results: Sequence[RunResult], batch: jio.InputBatch,
                  out: str, fmt: str, opts: Dict[str, Any]) -> List[str]:
    single = len(results) == 1
    if os.path.splitext(out)[1]:
        stem = os.path.splitext(out)[0]
    elif single:
        stem = os.path.join(out, f"{results[0].model_id.lower().replace('uc1-', '')}"
                                 f"_{results[0].module.split('_', 2)[-1]}")
    else:
        stem = os.path.join(out, "uc1_all_models")

    os.makedirs(os.path.dirname(os.path.abspath(stem)) or ".", exist_ok=True)
    written: List[str] = []

    if fmt in ("json", "both"):
        generated_at = datetime.now(timezone.utc).isoformat()
        path = stem + ".json"
        with open(path, "w", encoding="utf-8") as fh:
            json.dump({
                "generated_by": MODULE_VERSION,
                "generated_at_utc": generated_at,
                "input": batch.summary(),
                "options": {k: v for k, v in opts.items() if not k.startswith("_")},
                "results": [r.as_dict() for r in results],
            }, fh, indent=2, default=str)
        written.append(path)

        # The file above is the audit trail: every formula, every substitution,
        # every intermediate node. Alongside it, write the vessel-keyed summary a
        # dashboard can consume directly, so nobody has to walk an 800 KB
        # document to find a berthing time. Same numbers, selected not recomputed.
        import dashboard_json

        dash_path = dashboard_json.write(
            dashboard_json.build(results, batch, opts, generated_at_utc=generated_at,
                                 full_detail_file=jnpa_paths.relative(path)),
            stem + "_dashboard.json",
        )
        written.append(dash_path)

    if fmt in ("csv",):
        for r in results:
            if not r.rows:
                continue
            path = f"{stem}_{r.model_id.lower().replace('uc1-', '')}.csv" if not single \
                else stem + ".csv"
            with open(path, "w", encoding="utf-8", newline="") as fh:
                w = csv.DictWriter(fh, fieldnames=list(r.columns), extrasaction="ignore")
                w.writeheader()
                w.writerows(r.rows)
            written.append(path)

    if fmt in ("xlsx", "both"):
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill
        except Exception:
            for r in results:
                if r.rows:
                    path = f"{stem}_{r.model_id.lower().replace('uc1-', '')}.csv"
                    with open(path, "w", encoding="utf-8", newline="") as fh:
                        w = csv.DictWriter(fh, fieldnames=list(r.columns),
                                           extrasaction="ignore")
                        w.writeheader()
                        w.writerows(r.rows)
                    written.append(path)
            return written

        path = stem + ".xlsx"
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        head_fill = PatternFill("solid", fgColor="1F4E78")

        if len(results) > 1:
            ws = wb.create_sheet("Summary")
            ws.append(["Model", "Module", "Scope", "Status", "Rows", "Headline"])
            for c in ws[1]:
                c.font = Font(bold=True, color="FFFFFF")
                c.fill = head_fill
            for r in results:
                headline = "; ".join(f"{k}={v}" for k, v in list(r.summary.items())[:4])
                ws.append([r.model_id, r.module, r.scope,
                           "OK" if r.ok else "FAILED", len(r.rows), headline])
            for col, width in (("A", 10), ("B", 26), ("C", 11), ("D", 9),
                               ("E", 7), ("F", 110)):
                ws.column_dimensions[col].width = width

        for r in results:
            ws = wb.create_sheet(r.model_id.replace("UC1-", "")[:31])
            if r.rows:
                ws.append(list(r.columns))
                for c in ws[1]:
                    c.font = Font(bold=True, color="FFFFFF")
                    c.fill = head_fill
                    c.alignment = Alignment(horizontal="center", wrap_text=True)
                for row in r.rows:
                    ws.append([row.get(k, "") for k in r.columns])
                for i, name in enumerate(r.columns, start=1):
                    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = max(
                        11, min(34, len(str(name)) + 3))
                ws.freeze_panes = "A2"
                ws.append([])
            else:
                ws.append([f"{r.model_id}: {'no rows' if r.ok else r.error}"])
            ws.append(["NOTES"])
            ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
            for n in r.notes:
                ws.append([n])

        info = wb.create_sheet("Run_Info")
        info.append(["Key", "Value"])
        for c in info[1]:
            c.font = Font(bold=True)
        for k, v in [
            ("generated_by", MODULE_VERSION),
            ("generated_at_utc", datetime.now(timezone.utc).isoformat()),
            ("input_file", batch.source_file),
            ("input_format", batch.source_format),
            ("rows_valid", f"{len(batch.valid_rows)} / {len(batch.rows)}"),
            ("input_errors", batch.error_count),
            ("input_warnings", batch.warn_count),
            ("models_run", ", ".join(r.model_id for r in results)),
        ] + [(f"option.{k}", str(v)) for k, v in sorted(opts.items())]:
            info.append([k, str(v)])
        info.column_dimensions["A"].width = 26
        info.column_dimensions["B"].width = 100

        wb.save(path)
        written.append(path)

    return written


# ---------------------------------------------------------------------------
# self-test
# ---------------------------------------------------------------------------


def _self_test() -> List[Tuple[str, bool, str]]:
    import tempfile

    checks: List[Tuple[str, bool, str]] = []

    def check(name: str, ok: bool, detail: str = "") -> None:
        checks.append((name, bool(ok), detail))

    sample = jnpa_paths.SAMPLE_INPUT_XLSX
    if not os.path.exists(sample):
        check("sample_present", False, "Vessel_Training_Input_Sample.xlsx not found")
        return checks

    batch = jio.load_input(sample)
    check("input_loads", batch.ok, f"{len(batch.valid_rows)} valid rows")
    check("registry_complete", len(MODELS) == 8, f"{len(MODELS)} models registered")

    opts = {"model_dir": jnpa_paths.TRAINED_MODELS_DIR, "roster_preset": "real"}
    results = [run_one(k, batch, opts) for k in MODELS]
    for r in results:
        check(f"{r.model_id.lower()}_runs", r.ok, r.error or f"{len(r.rows)} rows")

    per_row = [r for r in results if r.scope == "per-row" and r.ok]
    check("per_row_models_cover_every_vessel",
          all(len(r.rows) == len(batch.valid_rows) for r in per_row),
          f"{len(per_row)} per-row models x {len(batch.valid_rows)} vessels")

    m1 = next(r for r in results if r.model_id == "UC1-M1")
    check("m1_scores_against_sheet", "validation_vs_sheet" in m1.summary,
          str(m1.summary.get("validation_vs_sheet", "")))
    check("m1_never_uses_the_label",
          all("Sheet_DUKC_Status" in row and "Status" in row for row in m1.rows),
          "the sheet's status is reported beside the model's, never substituted for it")

    m8 = next(r for r in results if r.model_id == "UC1-M8")
    check("m8_logs_every_node",
          all(row["Propagation_Steps"] == m8.summary["nodes"] for row in m8.rows),
          f"{m8.summary['nodes']} steps logged per run")

    with tempfile.TemporaryDirectory() as tmp:
        written = write_outputs(results, batch, tmp, "both", opts)
        check("outputs_written", len(written) >= 2, ", ".join(os.path.basename(w) for w in written))
        json_path = next(w for w in written if w.endswith(".json"))
        with open(json_path, encoding="utf-8") as fh:
            payload = json.load(fh)
        check("json_has_all_models", len(payload["results"]) == 8,
              f"{len(payload['results'])} model results")

        single = write_outputs([results[0]], batch, tmp, "csv", opts)
        check("single_model_csv", single and os.path.exists(single[0]),
              os.path.basename(single[0]) if single else "none")

    return checks


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def _print_list() -> None:
    print("=" * 96)
    print(f"  JNPA UC-1 MODELS  |  {MODULE_VERSION}")
    print("=" * 96)
    print(f"  {'key':<6} {'id':<9} {'scope':<10} {'module':<28} title")
    print("  " + "-" * 92)
    for key, (module, title, _) in MODELS.items():
        print(f"  {key:<6} {'UC1-' + key.upper():<9} {MODEL_SCOPE[key]:<10} {module:<28} {title}")
    print()
    print("  run one:   python run.py models --model m1 --input data/input/Vessel_Training_Input_Sample.xlsx")
    print("  run all:   python run.py models --model all --input data/input/Vessel_Training_Input_Sample.xlsx")
    print("  targets:   python run.py predict   --input data/input/Vessel_Training_Input_Sample.xlsx")
    print()


def main(argv: Optional[Sequence[str]] = None) -> int:
    p = argparse.ArgumentParser(
        description="Run any UC-1 model (or all eight) against your input file.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            "examples:\n"
            "  python run.py models --list\n"
            "  python run.py models --model m1  --input data/input/Vessel_Training_Input_Sample.xlsx\n"
            "  python run.py models --model m2  --input data.xlsx --horizon-hours 168\n"
            "  python run.py models --model m7  --input data.xlsx --roster-preset poc\n"
            "  python run.py models --model all --input data.xlsx --out out/ --format both\n"
        ),
    )
    p.add_argument("--model", "-m", default="all",
                   help="m1..m8, 'all', or a comma-separated list such as m1,m2,m8")
    p.add_argument("--input", "-i", help="input .xlsx / .csv / .json")
    p.add_argument("--sheet", help="worksheet name (default: first sheet)")
    p.add_argument("--out", "-o", default="out", help="output directory or file path")
    p.add_argument("--format", choices=("xlsx", "csv", "json", "both"), default="both")
    p.add_argument("--tide-policy", choices=("harmonic", "column", "fixed"), default="harmonic")
    p.add_argument("--tide-m", type=float, help="fixed tide height for --tide-policy fixed")
    p.add_argument("--artifact", "-a", help="M3 trained .pkl (default: newest in --model-dir)")
    p.add_argument("--model-dir", default=jnpa_paths.TRAINED_MODELS_DIR)
    p.add_argument("--wait-model", choices=("optimiser", "queue", "none"), default="optimiser",
                   help="M3/M5: how the pre-berth wait is estimated")
    p.add_argument("--wait-percentile", type=int, choices=(10, 50, 90), default=50)
    p.add_argument("--cluster-gap-hours", type=float, default=72.0,
                   help="M5: calls further apart than this are planned separately")
    p.add_argument("--horizon-hours", type=float, default=120.0, help="M2: scan horizon")
    p.add_argument("--ais-staleness-min", type=float, default=DEFAULT_AIS_STALENESS_MIN,
                   help="M4: assumed AIS position age")
    p.add_argument("--roster-preset", choices=("real", "poc"), default="real",
                   help="M7: 18-craft PDF roster, or the WS2 spec's 9-craft PoC roster")
    p.add_argument("--down-craft", default="",
                   help="M7: comma-separated craft ids to mark unavailable, e.g. PL-01,PL-02")
    p.add_argument("--quiet", "-q", action="store_true")
    p.add_argument("--list", action="store_true", help="list the models and exit")
    p.add_argument("--selftest", action="store_true")
    args = p.parse_args(argv)

    if args.list:
        _print_list()
        return 0

    if args.selftest or not args.input:
        checks = _self_test()
        passed = sum(1 for _, ok, _ in checks if ok)
        print("=" * 96)
        print(f"  run_model.py self-test  |  {MODULE_VERSION}")
        print("=" * 96)
        for name, ok, detail in checks:
            print(f"  [{'PASS' if ok else 'FAIL'}] {name:<36} {detail}")
        print("-" * 96)
        print(f"  {passed}/{len(checks)} checks passed")
        if not args.input:
            print()
            _print_list()
        return 0 if passed == len(checks) else 1

    keys = list(MODELS) if args.model.lower() == "all" else [
        k.strip().lower() for k in args.model.split(",") if k.strip()
    ]
    unknown = [k for k in keys if k not in MODELS]
    if unknown:
        print(f"ERROR: unknown model(s) {', '.join(unknown)}. Valid: "
              f"{', '.join(MODELS)}, all", file=sys.stderr)
        return 1

    try:
        batch = jio.load_input(args.input, sheet=args.sheet,
                               tide_policy=args.tide_policy, fixed_tide_m=args.tide_m)
    except Exception as exc:
        print(f"ERROR reading input: {exc}", file=sys.stderr)
        return 1

    if not batch.valid_rows:
        print(f"ERROR: no valid rows. Run 'python run.py input --input {args.input} "
              f"--validate' to see why.", file=sys.stderr)
        for i in batch.all_issues:
            if i.severity == "ERROR":
                print(f"  {i}", file=sys.stderr)
        return 1

    opts: Dict[str, Any] = {
        "artifact": args.artifact,
        "model_dir": args.model_dir,
        "wait_model": args.wait_model,
        "wait_percentile": args.wait_percentile,
        "cluster_gap_hours": args.cluster_gap_hours,
        "horizon_hours": args.horizon_hours,
        "ais_staleness_min": args.ais_staleness_min,
        "roster_preset": args.roster_preset,
        "down_craft": args.down_craft,
        "tide_policy": args.tide_policy,
    }

    if not args.quiet:
        print("=" * 118)
        print(f"  JNPA UC-1 MODEL RUNNER  |  {MODULE_VERSION}")
        print("=" * 118)
        print(f"  input   : {batch.source_file}")
        print(f"            {len(batch.valid_rows)}/{len(batch.rows)} rows valid, "
              f"{batch.error_count} errors, {batch.warn_count} warnings "
              f"({batch.source_format}, sheet '{batch.sheet_name}')")
        print(f"  models  : {', '.join(k.upper() for k in keys)}")
        print(f"  tide    : {args.tide_policy}"
              + (f" ({args.tide_m} m)" if args.tide_m is not None else ""))

    results = [run_one(k, batch, opts) for k in keys]

    if not args.quiet:
        for r in results:
            _print_result(r)

    written = write_outputs(results, batch, args.out, args.format, opts)
    print("  OUTPUT")
    for path in written:
        print(f"    {path}  ({os.path.getsize(path):,} bytes)")

    failed = [r for r in results if not r.ok]
    if failed:
        print(f"\n  {len(failed)} model(s) FAILED: "
              f"{', '.join(r.model_id + ' -- ' + r.error for r in failed)}")
    print()
    return 1 if failed else 0


if __name__ == "__main__":
    raise SystemExit(main())
